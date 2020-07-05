import csv
import math
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
states = ['TT','AN','AP','AR','AS','BR','CH','CT','DN','DD','DL','GA','GJ','HR','HP','JK','JH','KA','KL','LA','LD','MP','MH','MN','ML','MZ','NL','OR','PY','PB','RJ','SK','TN','TG','TR','UP','UT','WB','UN']
read_file = pd.read_csv (r'C:\Users\ddd\Documents\DIC Mech\Coronavirus Dynamic Reporting\state_wise_daily.csv', low_memory=False)
read_file.to_excel (r'C:\Users\ddd\Documents\DIC Mech\Coronavirus Dynamic Reporting\state_wise_daily.xlsx', index = None, header=True)
wb = load_workbook(r'C:\Users\ddd\Documents\DIC Mech\Coronavirus Dynamic Reporting\state_wise_daily.xlsx')
ws_main = wb.worksheets[0]
sheet = wb.active
mr = ws_main.max_row
mc = ws_main.max_column
for i in range (3, mc + 1):
    temp_wb = Workbook()
    ws = temp_wb.active
    iter1 = 2
    ws.cell(row = 1, column = 1).value = 'Date'
    ws.cell(row = 1, column = 2).value = 'Confirmed Cases'
    ws.cell(row = 1, column = 3).value = 'Recovered Cases'
    ws.cell(row = 1, column = 4).value = 'No. of Deaths'
    ws.cell(row = 1, column = 5).value = 'Cumulative Confirmed'
    ws.cell(row = 1, column = 6).value = 'Cumulative Recovered'
    ws.cell(row = 1, column = 7).value = 'Cumulative Deaths'
    ws.cell(row = 1, column = 8).value = 'Recovery Rate'
    ws.cell(row = 1, column = 9).value = 'Death Rate'
    #ws.cell(row = 1, column = 10).value = 'Doubling Rate'
    for j in range (2, mr + 1):
        if(j%3==2):
            date = ws_main.cell(row = j, column = 1)
            ws.cell(row = iter1, column = 1).value = date.value
        data = ws_main.cell(row = j, column = i)
        if(ws_main.cell(row = j, column = 2).value=='Confirmed'):
            ws.cell(row = iter1, column = 2).value = data.value
            if(iter1==2):
                ws.cell(row = iter1, column = 5).value = ws.cell(row = iter1, column = 2).value
            else:
                ws.cell(row = iter1, column = 5).value = ws.cell(row = iter1, column = 2).value + ws.cell(row = iter1-1, column = 5).value
        elif(ws_main.cell(row = j, column = 2).value=='Recovered'):
            ws.cell(row = iter1, column = 3).value = data.value
            if(iter1==2):
                ws.cell(row = iter1, column = 6).value = ws.cell(row = iter1, column = 3).value
            else:
                ws.cell(row = iter1, column = 6).value = ws.cell(row = iter1, column = 3).value + ws.cell(row = iter1-1, column = 6).value
            if(ws.cell(row = iter1, column = 5).value !=0):
                ws.cell(row = iter1, column = 8).value = round((ws.cell(row = iter1, column = 6).value/ws.cell(row = iter1, column = 5).value)*100,2)
            else:
                ws.cell(row = iter1, column = 8).value = 0
        elif(ws_main.cell(row = j, column = 2).value=='Deceased'):
            ws.cell(row = iter1, column = 4).value = data.value
            if(iter1==2):
                ws.cell(row = iter1, column = 7).value = ws.cell(row = iter1, column = 4).value
            else:
                ws.cell(row = iter1, column = 7).value = ws.cell(row = iter1, column = 4).value + ws.cell(row = iter1-1, column = 7).value
            if(ws.cell(row = iter1, column = 5).value !=0):
                ws.cell(row = iter1, column = 9).value = round((ws.cell(row = iter1, column = 7).value/ws.cell(row = iter1, column = 5).value)*100,2)
            else:
                ws.cell(row = iter1, column = 9).value = 0
            iter1+=1
    temp_wb.save(str(i-2)+states[i-3]+".xlsx")
    print(states[i-3])
wb.save(r'C:\Users\ddd\Documents\DIC Mech\Coronavirus Dynamic Reporting\state_wise_daily.xlsx')
